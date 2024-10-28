import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 定义合并和处理 Excel 文件的函数
def merge_excel_files(folder_path, output_path):
    all_data_sheet1 = pd.DataFrame()
    all_data_sheet2 = pd.DataFrame()

    for filename in os.listdir(folder_path):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(folder_path, filename)
            excel_data = pd.read_excel(file_path, sheet_name=None)
            sheet_names = list(excel_data.keys())
            
            if len(sheet_names) >= 2:
                data_sheet1 = excel_data[sheet_names[0]]
                data_sheet2 = excel_data[sheet_names[1]]
                all_data_sheet1 = pd.concat([all_data_sheet1, data_sheet1], ignore_index=True)
                all_data_sheet2 = pd.concat([all_data_sheet2, data_sheet2], ignore_index=True)
            else:
                print(f"文件 {filename} 中没有足够的工作表。")

    def contains_chinese(s):
        for char in str(s):
            if '\u4e00' <= char <= '\u9fff':
                return True
        return False

    all_data_sheet1 = all_data_sheet1[~((all_data_sheet1.index >= 2) & all_data_sheet1.iloc[:, 0].apply(contains_chinese))]
    all_data_sheet1 = all_data_sheet1.drop_duplicates()
    all_data_sheet2 = all_data_sheet2[~((all_data_sheet2.index >= 2) & all_data_sheet2.iloc[:, 0].apply(contains_chinese))]
    all_data_sheet2 = all_data_sheet2.drop_duplicates()

    all_data_sheet1 = all_data_sheet1.dropna(how='all', axis=1)
    all_data_sheet2 = all_data_sheet2.dropna(how='all', axis=1)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        all_data_sheet1.to_excel(writer, sheet_name='Sheet1', index=False)
        all_data_sheet2.to_excel(writer, sheet_name='Sheet2', index=False)

    wb = load_workbook(output_path)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            sheet.column_dimensions[column_letter].width = max_length + 2
    wb.save(output_path)
    print("所有文件已成功合并并保存到", output_path)
    messagebox.showinfo("完成", "Excel 文件合并和处理已完成！")

# 定义选择文件夹路径的函数
def select_folder():
    folder_selected = filedialog.askdirectory()
    folder_entry.delete(0, tk.END)
    folder_entry.insert(0, folder_selected)

# 定义保存文件的路径选择函数
def select_output_file():
    output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    output_entry.delete(0, tk.END)
    output_entry.insert(0, output_file)

# 定义开始处理的函数
def start_processing():
    folder_path = folder_entry.get()
    output_path = output_entry.get()
    if not folder_path or not output_path:
        messagebox.showwarning("路径错误", "请选择有效的文件夹和输出文件路径")
        return
    merge_excel_files(folder_path, output_path)

# 创建主窗口
root = tk.Tk()
root.title("校级信息员|表格文件合并工具")
root.geometry("700x250")
root.configure(bg="#fdfdfd")

# 添加标题 - 普通样式
title_label = tk.Label(root, text="注意：每个表格文件需要相同格式", font=12, bg="#fdfdfd")
title_label.pack(pady=10)

# 添加文件夹选择框
frame = tk.Frame(root, bg="#fdfdfd")
frame.pack(pady=5)

tk.Label(frame, text="选择文件夹路径:", bg="#fdfdfd", fg="#333333").grid(row=0, column=0, padx=10, pady=5, sticky="e")
folder_entry = tk.Entry(frame, width=45)
folder_entry.grid(row=0, column=1, padx=10, pady=5)
folder_button = ttk.Button(frame, text="浏览", command=select_folder)
folder_button.grid(row=0, column=2, padx=5, pady=5)

# 添加输出文件路径选择框
tk.Label(frame, text="输出文件的路径:", bg="#fdfdfd", fg="#333333").grid(row=1, column=0, padx=10, pady=5, sticky="e")
output_entry = tk.Entry(frame, width=45)
output_entry.grid(row=1, column=1, padx=10, pady=5)
output_button = ttk.Button(frame, text="浏览", command=select_output_file)
output_button.grid(row=1, column=2, padx=5, pady=5)

# 添加开始按钮
start_button = ttk.Button(root, text="开始处理", command=start_processing, width=20)
start_button.pack(pady=15)

# 设置按钮样式
style = ttk.Style()
style.configure("TButton", padding=5)

# 运行主循环
root.mainloop()
